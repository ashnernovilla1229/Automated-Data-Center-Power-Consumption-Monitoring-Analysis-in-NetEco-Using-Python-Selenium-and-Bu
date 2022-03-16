# -*- coding: utf-8 -*-
"""
Created on Tue Feb  8 19:25:37 2022

@author: Openlab
"""

# pip install lxml bs4 html5lib  #To use webscrapper for table

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from selenium import webdriver

import ssl
import csv  
from datetime import datetime, timedelta
import os
import time
import pandas as pd
import win32com.client


def NetEco_Webloading():
    ssl._create_default_https_context = ssl._create_unverified_context

    global driver #Making driver as a global variable
    
    driver = webdriver.Chrome(executable_path=r'C:\Users\Openlab\Documents\SeleniumDrivers\chromedriver.exe') 
    driver.implicitly_wait(60)

    TestName = "NetEco_WebLoading"
    Time_BeforeTest = datetime.now() #Time log the start of the test after opening the browser
            
    driver.get("https://10.1.1.15:31943/") #Test Link

    try:
        driver.find_element_by_id("details-button").click()  #if a google warning appear run this code
        driver.find_element_by_id("proceed-link").click()
    except: 
        print("Verification Done") #If no error pass to exception

    Time_AfterTest = datetime.now() #Time log after pressing the submit button

    result = "Pass"
    TimeDiff = abs(Time_AfterTest - Time_BeforeTest) #Difference of before and after the test

    print(TestName, TimeDiff, result)   

def NetEco_Login():
    for retry in range(3):
        retry = retry+1    
        try:
             TestName = "NetEco_Login" # Title of the test
             Time_BeforeTest = datetime.now() #Time log the start of the test after opening the browser
             driver.find_element_by_xpath("//input[@id='txf_username']").send_keys("admin")    
             driver.find_element_by_xpath("//input[@id='txf_imtinfo']").send_keys("Huawei123")
             driver.find_element_by_xpath("//div[@id='btn_submit']").click()
             
             Time_AfterTest = datetime.now() #Time log after pressing the submit button
             TimeDiff = abs(Time_AfterTest - Time_BeforeTest) #Difference of before and after the test
             result = "Pass"
             # logtest(TestName , Time_BeforeTest, Time_AfterTest, TimeDiff, result) #Datalog to the csv sheet
             
             print(TestName, TimeDiff, result)
             
             break
             
        except:
            if retry > 2:
                print("The Test Fails in Log-in Please Re Run the Code", 0/0)
            else:
                print("Number of test done is: ", retry)
                continue

def License_Near_Expiry():
    for retry in range(3):
        retry = retry+1    
        try:
             "License_Near_Expiration_Exception" # Title of the test
             driver.find_element_by_id('license_regist_15_1').click()
             
             break
             
        except:
            if retry > 2:
                break
            else:
                print("Number of test done is: ", retry)
                continue

    
def Historical_Report():
    for retry in range(3):
        retry = retry+1
        try:
             TestName = "Historical_Report" # Title of the test
             Time_BeforeTest = datetime.now() #Time log the start of the test after opening the browser
    
             driver.find_element_by_xpath("//body/div[@id='globe_header']/div[1]/div[2]/ul[1]/li[2]").click()
             driver.find_element_by_xpath("//a[@id='menu.com.huawei.neteco.report.ui.menu.history.neteco']").click()
             
             driver.find_element_by_xpath("//div[@id='pm_historyData_timeRangeSelect']").click()
             driver.find_element_by_xpath("//label[contains(text(),'Customize')]").click()
             
             Time_AfterTest = datetime.now() #Time log after pressing the submit button
             TimeDiff = abs(Time_AfterTest - Time_BeforeTest) #Difference of before and after the test
             result = "Pass"
             
             print(TestName, TimeDiff, result) 
             break
            
        except:
            if retry > 2:
                print("The Test Fails in Historical Report Please Re Run the Code", 0/0)
            else:
                print("Number of test done is: ", retry)
                continue

         

def Set_Duration_Data():
    for retry in range(3):
        retry = retry+1
        try:
             TestName = "Set_Date_Time" # Title of the test
             Time_BeforeTest = datetime.now() #Time log the start of the test after opening the browser
             
             date_diff = datetime.now() + timedelta(days=-1) 
             #date_diff = datetime.now() + timedelta(days=5, hours=-5)
    
             driver.find_element_by_xpath("//div[@id='pm_historyData_timeRangeSelect']").click()
             driver.find_element_by_xpath("//label[contains(text(),'Customize')]").click()
             
             driver.find_element_by_xpath("//input[@id='pm_pmcommon_fromDate_value']").clear()  #Clear the old date "From"
             driver.find_element_by_xpath("//input[@id='pm_pmcommon_toDate_value']").clear() #Clear the old date "To"
             
             #Comment this if specific date is required as this take the data from yesterday 0:00 to 23:00
             driver.find_element_by_xpath("//input[@id='pm_pmcommon_fromDate_value']").send_keys(date_diff.strftime("%Y-%m-%d 00:00"))
             driver.find_element_by_xpath("//input[@id='pm_pmcommon_toDate_value']").send_keys(date_diff.strftime("%Y-%m-%d 23:00")) 
    
             #Uncomment code below if need to to run a specific date 
             # driver.find_element_by_xpath("//input[@id='pm_pmcommon_fromDate_value']").send_keys("2022-03-14 00:00")
             # driver.find_element_by_xpath("//input[@id='pm_pmcommon_toDate_value']").send_keys("2022-03-14 23:00") 
    
              
             Time_AfterTest = datetime.now() #Time log after pressing the submit button
             TimeDiff = abs(Time_AfterTest - Time_BeforeTest) #Difference of before and after the test
             result = "Pass"
             
             print(TestName, TimeDiff, result) 
             break
             
        except:
            if retry > 2:
                print("The Test Fails in Set Duration - Please Re Run the Code", 0/0)
            else:
                print("Number of test done is: ", retry)
                continue

def Expand_Openlab_DC_Monitoring():
    for retry in range(3):
        retry = retry+1    
        try:
             TestName = "Openlab_DC_Monitoring_Expansion" # Title of the test
             Time_BeforeTest = datetime.now() #Time log the start of the test after opening the browser
            
             driver.find_element_by_id("pm_historyData_neTree_element_2_switch").click()
             driver.find_element_by_id("pm_historyData_neTree_element_3_switch").click()
             driver.find_element_by_id("pm_historyData_neTree_element_4_switch").click()
             
             Time_AfterTest = datetime.now() #Time log after pressing the submit button
             TimeDiff = abs(Time_AfterTest - Time_BeforeTest) #Difference of before and after the test
             result = "Pass"
             
             print(TestName, TimeDiff, result) 
             break
             
        except:
            if retry > 2:
                print("The Test Fails in Expand_Openlab_DC_Monitoring - Please Re Run the Code", 0/0)
            else:
                print("Number of test done is: ", retry)
                continue  

def UPS_001_40Mod_ACEnergy():
    for retry in range(3):
        retry = retry+1    
        try:
         TestName = "UPS_001_40Mod_ACEnergy_Expansion" # Title of the test
         Time_BeforeTest = datetime.now() #Time log the start of the test after opening the browser
        
         driver.find_element_by_id("pm_historyData_neTree_element_6_switch").click()
         driver.find_element_by_id("pm_historyData_neTree_element_68_check").click()
         
         driver.find_element_by_xpath("//div[@id='pm_historyData_countergroup']").click() #This is the Counter Group
         driver.find_element_by_xpath("//label[contains(text(),'Air Conditioner Branch - Energy Information Measur')]").click()
         driver.find_element_by_xpath("//div[@id='pm_historyData_query']").click()  #This is the query button
         
         driver.implicitly_wait(10)  #Conditional wait if the data is existing for 10 sec - if not existing go to exception result
         driver.find_element_by_id("resultTable_tableScrollContainer")
         
         driver.find_element_by_xpath("//span[@id='resultTable_pagination_widget_input']").click()
         driver.find_element_by_css_selector("#resultTable_pagination_widget_page_size_50").click()
         
         time.sleep(10)  #Pause for a while to give sometime for the api to send the data 
         
         driver.implicitly_wait(10)  #Conditional wait if the data is existing for 10 sec - if not existing go to exception result
         driver.find_element_by_id("resultTable_tree_table_scroll")
         
         df=pd.read_html(driver.find_element_by_id("resultTable_table").get_attribute('outerHTML'))[0]
         df2=pd.read_html(driver.find_element_by_id("resultTable_tree_table_scroll").get_attribute('outerHTML'))[0]
         df_tbl_name = list(df.columns)

         global UPS_001_40Mod_AC_DF  #Converted to global because we want to store the data into DataBase Later
         UPS_001_40Mod_AC_DF = df2.set_axis(df_tbl_name, axis=1, inplace=False)
         
         
         del(df,df2,df_tbl_name)
     
         driver.refresh()
     
         Time_AfterTest = datetime.now() #Time log after pressing the submit button
         TimeDiff = abs(Time_AfterTest - Time_BeforeTest) #Difference of before and after the test
         result = "Pass"
         
         print(TestName, TimeDiff, result) 
         
         break
     
        except:
            if retry > 2:
                print("The Test Fails in UPS_001_40Mod_ACEnergy - Please Re Run the Code", 0/0)
            else:
                print("Number of test done is: ", retry)
                continue  

def UPS_001_40Mod_ITB1_EnergyInfo():
    for retry in range(3):
        retry = retry+1        
        try:
             TestName = "UPS_001_40Mod_ITB1_EnergyInfo" # Title of the test
             Time_BeforeTest = datetime.now() #Time log the start of the test after opening the browser
            
             driver.find_element_by_id("pm_historyData_neTree_element_6_switch").click()
             driver.find_element_by_id("pm_historyData_neTree_element_68_check").click()
             
             driver.find_element_by_xpath("//div[@id='pm_historyData_countergroup']").click() #This is the Counter Group
             driver.find_element_by_xpath("//label[contains(text(),'IT Branch 1 - Energy Information Measurement by Ho')]").click()
             driver.find_element_by_xpath("//div[@id='pm_historyData_query']").click()  #This is the query button
             
             driver.implicitly_wait(10)  #Conditional wait if the data is existing for 10 sec - if not existing go to exception result
             driver.find_element_by_id("resultTable_tableScrollContainer")
             
             driver.find_element_by_xpath("//span[@id='resultTable_pagination_widget_input']").click()
             driver.find_element_by_css_selector("#resultTable_pagination_widget_page_size_50").click()
             
             time.sleep(10)  #Pause for a while to give sometime for the api to send the data 
             
             driver.implicitly_wait(10)  #Conditional wait if the data is existing for 10 sec - if not existing go to exception result
             driver.find_element_by_id("resultTable_tree_table_scroll")
             
             df=pd.read_html(driver.find_element_by_id("resultTable_table").get_attribute('outerHTML'))[0]
             df2=pd.read_html(driver.find_element_by_id("resultTable_tree_table_scroll").get_attribute('outerHTML'))[0]
             df_tbl_name = list(df.columns)
    
             global UPS_001_40Mod_ITB1_EnergyInfo_DF  #Converted to global because we want to store the data into DataBase Later
             UPS_001_40Mod_ITB1_EnergyInfo_DF = df2.set_axis(df_tbl_name, axis=1, inplace=False)
             
             del(df,df2,df_tbl_name)
         
             driver.refresh()
         
             Time_AfterTest = datetime.now() #Time log after pressing the submit button
             TimeDiff = abs(Time_AfterTest - Time_BeforeTest) #Difference of before and after the test
             result = "Pass"
             
             print(TestName, TimeDiff, result)  
             break
    
        except:
            if retry > 2:
                print("The Test Fails in UPS_001_40Mod_ITB1_EnergyInfo - Please Re Run the Code", 0/0)
            else:
                print("Number of test done is: ", retry)
                continue  

     
def UPS_001_40Mod_ITB2_EnergyInfo():
    for retry in range(3):
        retry = retry+1 
    
        try:
             TestName = "UPS_001_40Mod_ITB2_EnergyInfo" # Title of the test
             Time_BeforeTest = datetime.now() #Time log the start of the test after opening the browser
            
             driver.find_element_by_id("pm_historyData_neTree_element_6_switch").click()
             driver.find_element_by_id("pm_historyData_neTree_element_68_check").click()
             
             driver.find_element_by_id("pm_historyData_countergroup").click() 
             driver.find_element_by_id("pm_historyData_countergroup_select_input").clear()
             driver.find_element_by_id("pm_historyData_countergroup_select_input").send_keys("IT Branch 2 - Energy Information Measurement by Hour")
        
             driver.find_element_by_xpath("//div[@id='pm_historyData_query']").click()  #This is the query button
             
             driver.implicitly_wait(10)  #Conditional wait if the data is existing for 10 sec - if not existing go to exception result
             driver.find_element_by_id("resultTable_tableScrollContainer")
             
             driver.find_element_by_xpath("//span[@id='resultTable_pagination_widget_input']").click()
             driver.find_element_by_css_selector("#resultTable_pagination_widget_page_size_50").click()
             
             time.sleep(10)  #Pause for a while to give sometime for the api to send the data 
    
             driver.implicitly_wait(10)  #Conditional wait if the data is existing for 10 sec - if not existing go to exception result
             driver.find_element_by_id("resultTable_tree_table_scroll")
             
             df=pd.read_html(driver.find_element_by_id("resultTable_table").get_attribute('outerHTML'))[0]
             df2=pd.read_html(driver.find_element_by_id("resultTable_tree_table_scroll").get_attribute('outerHTML'))[0]
             df_tbl_name = list(df.columns)
    
             global UPS_001_40Mod_ITB2_EnergyInfo_DF  #Converted to global because we want to store the data into DataBase Later
             UPS_001_40Mod_ITB2_EnergyInfo_DF = df2.set_axis(df_tbl_name, axis=1, inplace=False)
             
             del(df,df2,df_tbl_name)
         
             driver.refresh()
         
             Time_AfterTest = datetime.now() #Time log after pressing the submit button
             TimeDiff = abs(Time_AfterTest - Time_BeforeTest) #Difference of before and after the test
             result = "Pass"
             
             print(TestName, TimeDiff, result) 
             break
    
        except:
            if retry > 2:
                print("The Test Fails in UPS_001_40Mod_ITB2_EnergyInfo - Please Re Run the Code", 0/0)
            else:
                print("Number of test done is: ", retry)
                continue  

      

def UPS_02_40Mod_ITB1_EnergyInfo():
    for retry in range(3):
        retry = retry+1 
        try:
             TestName = "UPS_02_40Mod_ITB1_EnergyInfo" # Title of the test
             Time_BeforeTest = datetime.now() #Time log the start of the test after opening the browser
            
             driver.find_element_by_id("pm_historyData_neTree_element_6_switch").click()
             driver.find_element_by_id("pm_historyData_neTree_element_69_check").click()
             
             driver.find_element_by_xpath("//div[@id='pm_historyData_countergroup']").click() #This is the Counter Group
             driver.find_element_by_xpath("//label[contains(text(),'IT Branch 1 - Energy Information Measurement by Ho')]").click()
             driver.find_element_by_xpath("//div[@id='pm_historyData_query']").click()  #This is the query button
         
             driver.implicitly_wait(10)  #Conditional wait if the data is existing for 10 sec - if not existing go to exception result
             driver.find_element_by_id("resultTable_tableScrollContainer")
             
             driver.find_element_by_xpath("//span[@id='resultTable_pagination_widget_input']").click()
             driver.find_element_by_css_selector("#resultTable_pagination_widget_page_size_50").click()
             
             time.sleep(10)  #Pause for a while to give sometime for the api to send the data 
             
             driver.implicitly_wait(10)  #Conditional wait if the data is existing for 10 sec - if not existing go to exception result
             driver.find_element_by_id("resultTable_tree_table_scroll")
             
             df=pd.read_html(driver.find_element_by_id("resultTable_table").get_attribute('outerHTML'))[0]
             df2=pd.read_html(driver.find_element_by_id("resultTable_tree_table_scroll").get_attribute('outerHTML'))[0]
             df_tbl_name = list(df.columns)
    
             global UPS_02_40Mod_ITB1_EnergyInfo_DF  #Converted to global because we want to store the data into DataBase Later
             UPS_02_40Mod_ITB1_EnergyInfo_DF = df2.set_axis(df_tbl_name, axis=1, inplace=False)
             
             del(df,df2,df_tbl_name)
         
             driver.refresh()
         
             Time_AfterTest = datetime.now() #Time log after pressing the submit button
             TimeDiff = abs(Time_AfterTest - Time_BeforeTest) #Difference of before and after the test
             result = "Pass"
             
             print(TestName, TimeDiff, result)  
             
             break
    
        except:
            if retry > 2:
                print("The Test Fails in UPS_02_40Mod_ITB1_EnergyInfo - Please Re Run the Code", 0/0)
            else:
                print("Number of test done is: ", retry)
                continue  

     
def UPS_02_40Mod_ITB2_EnergyInfo():
    for retry in range(3):
        retry = retry+1     
        try:
             TestName = "UPS_02_40Mod_ITB2_EnergyInfo" # Title of the test
             Time_BeforeTest = datetime.now() #Time log the start of the test after opening the browser
            
             driver.find_element_by_id("pm_historyData_neTree_element_6_switch").click()
             driver.find_element_by_id("pm_historyData_neTree_element_69_check").click()
             
             driver.find_element_by_id("pm_historyData_countergroup").click() 
             driver.find_element_by_id("pm_historyData_countergroup_select_input").clear()
             driver.find_element_by_id("pm_historyData_countergroup_select_input").send_keys("IT Branch 2 - Energy Information Measurement by Hour")
        
             driver.find_element_by_xpath("//div[@id='pm_historyData_query']").click()  #This is the query button
             
             driver.implicitly_wait(10)  #Conditional wait if the data is existing for 10 sec - if not existing go to exception result
             driver.find_element_by_id("resultTable_tableScrollContainer")
             
             driver.find_element_by_xpath("//span[@id='resultTable_pagination_widget_input']").click()
             driver.find_element_by_css_selector("#resultTable_pagination_widget_page_size_50").click()
    
             time.sleep(10)  #Pause for a while to give sometime for the api to send the data 
    
             driver.implicitly_wait(10)  #Conditional wait if the data is existing for 10 sec - if not existing go to exception result
             driver.find_element_by_id("resultTable_tree_table_scroll")
             
             df=pd.read_html(driver.find_element_by_id("resultTable_table").get_attribute('outerHTML'))[0]
             df2=pd.read_html(driver.find_element_by_id("resultTable_tree_table_scroll").get_attribute('outerHTML'))[0]
             df_tbl_name = list(df.columns)
    
             global UPS_02_40Mod_ITB2_EnergyInfo_DF  #Converted to global because we want to store the data into DataBase Later
             UPS_02_40Mod_ITB2_EnergyInfo_DF = df2.set_axis(df_tbl_name, axis=1, inplace=False)
             
             del(df,df2,df_tbl_name)
         
             driver.refresh()
         
             Time_AfterTest = datetime.now() #Time log after pressing the submit button
             TimeDiff = abs(Time_AfterTest - Time_BeforeTest) #Difference of before and after the test
             result = "Pass"
             print(TestName, TimeDiff, result) 
             
             break
    
        except:
            if retry > 2:
                print("The Test Fails in UPS_02_40Mod_ITB2_EnergyInfo - Please Re Run the Code", 0/0)
            else:
                print("Number of test done is: ", retry)
                continue  

      

def UPS_18Mod_ACEnergy():
    for retry in range(3):
        retry = retry+1  
        try:
             TestName = "UPS_18Mod_ACEnergy" # Title of the test
             Time_BeforeTest = datetime.now() #Time log the start of the test after opening the browser
            
             driver.find_element_by_id("pm_historyData_neTree_element_5_switch").click()
             driver.find_element_by_id("pm_historyData_neTree_element_35_check").click()
             
             driver.find_element_by_xpath("//div[@id='pm_historyData_countergroup']").click() #This is the Counter Group
             driver.find_element_by_xpath("//label[contains(text(),'Air Conditioner Branch - Energy Information Measur')]").click()
             driver.find_element_by_xpath("//div[@id='pm_historyData_query']").click()  #This is the query button
             
             driver.implicitly_wait(10)  #Conditional wait if the data is existing for 10 sec - if not existing go to exception result
             driver.find_element_by_id("resultTable_tableScrollContainer")
             
             driver.find_element_by_xpath("//span[@id='resultTable_pagination_widget_input']").click()
             driver.find_element_by_css_selector("#resultTable_pagination_widget_page_size_50").click()
             
             time.sleep(10)  #Pause for a while to give sometime for the api to send the data 
        
             driver.implicitly_wait(10)  #Conditional wait if the data is existing for 10 sec - if not existing go to exception result
             driver.find_element_by_id("resultTable_tree_table_scroll")
             
             df=pd.read_html(driver.find_element_by_id("resultTable_table").get_attribute('outerHTML'))[0]
             df2=pd.read_html(driver.find_element_by_id("resultTable_tree_table_scroll").get_attribute('outerHTML'))[0]
             df_tbl_name = list(df.columns)
        
             global UPS_18Mod_ACEnergy_DF  #Converted to global because we want to store the data into DataBase Later
             UPS_18Mod_ACEnergy_DF = df2.set_axis(df_tbl_name, axis=1, inplace=False)
             
             del(df,df2,df_tbl_name)
         
             driver.refresh()
         
             Time_AfterTest = datetime.now() #Time log after pressing the submit button
             TimeDiff = abs(Time_AfterTest - Time_BeforeTest) #Difference of before and after the test
             result = "Pass"
             
             print(TestName, TimeDiff, result)
             
             break
        
        except:
            if retry > 2:
                print("The Test Fails in UPS_18Mod_ACEnergy - Please Re Run the Code", 0/0)
            else:
                print("Number of test done is: ", retry)
                continue  

     

def UPS_18Mod_ITB1_EnergyInfo():
    for retry in range(3):
        retry = retry+1      
        try:
             TestName = "UPS_18Mod_ITB1_EnergyInfo" # Title of the test
             Time_BeforeTest = datetime.now() #Time log the start of the test after opening the browser
            
             driver.find_element_by_id("pm_historyData_neTree_element_5_switch").click()
             driver.find_element_by_id("pm_historyData_neTree_element_35_check").click()
             
             driver.find_element_by_xpath("//div[@id='pm_historyData_countergroup']").click() #This is the Counter Group
             driver.find_element_by_xpath("//label[contains(text(),'IT Branch 1 - Energy Information Measurement by Ho')]").click()
             driver.find_element_by_xpath("//div[@id='pm_historyData_query']").click()  #This is the query button
             
             driver.implicitly_wait(10)  #Conditional wait if the data is existing for 10 sec - if not existing go to exception result
             driver.find_element_by_id("resultTable_tableScrollContainer")
             
             driver.find_element_by_xpath("//span[@id='resultTable_pagination_widget_input']").click()
             driver.find_element_by_css_selector("#resultTable_pagination_widget_page_size_50").click()
    
             time.sleep(10)  #Pause for a while to give sometime for the api to send the data 
    
             driver.implicitly_wait(10)  #Conditional wait if the data is existing for 10 sec - if not existing go to exception result
             driver.find_element_by_id("resultTable_tree_table_scroll")
             
             df=pd.read_html(driver.find_element_by_id("resultTable_table").get_attribute('outerHTML'))[0]
             df2=pd.read_html(driver.find_element_by_id("resultTable_tree_table_scroll").get_attribute('outerHTML'))[0]
             df_tbl_name = list(df.columns)
    
             global UPS_18Mod_ITB1_EnergyInfo_DF  #Converted to global because we want to store the data into DataBase Later
             UPS_18Mod_ITB1_EnergyInfo_DF = df2.set_axis(df_tbl_name, axis=1, inplace=False)
             
             del(df,df2,df_tbl_name)
         
             driver.refresh()
         
             Time_AfterTest = datetime.now() #Time log after pressing the submit button
             TimeDiff = abs(Time_AfterTest - Time_BeforeTest) #Difference of before and after the test
             result = "Pass"
             
             print(TestName, TimeDiff, result)
             
             break
    
        except:
            if retry > 2:
                print("The Test Fails in UPS_18Mod_ITB1_EnergyInfo - Please Re Run the Code", 0/0)
            else:
                print("Number of test done is: ", retry)
                continue  
       

def UPS_18Mod_ITB2_EnergyInfo():
    for retry in range(3):
        retry = retry+1  
        try:
             TestName = "UPS_18Mod_ITB2_EnergyInfo" # Title of the test
             Time_BeforeTest = datetime.now() #Time log the start of the test after opening the browser
            
             driver.find_element_by_id("pm_historyData_neTree_element_5_switch").click()
             driver.find_element_by_id("pm_historyData_neTree_element_35_check").click()
             
             driver.find_element_by_id("pm_historyData_countergroup").click() 
             driver.find_element_by_id("pm_historyData_countergroup_select_input").clear()
             driver.find_element_by_id("pm_historyData_countergroup_select_input").send_keys("IT Branch 2 - Energy Information Measurement by Hour")
        
             driver.find_element_by_xpath("//div[@id='pm_historyData_query']").click()  #This is the query button
             
             driver.implicitly_wait(10)  #Conditional wait if the data is existing for 10 sec - if not existing go to exception result
             driver.find_element_by_id("resultTable_tableScrollContainer")
             
             driver.find_element_by_xpath("//span[@id='resultTable_pagination_widget_input']").click()
             driver.find_element_by_css_selector("#resultTable_pagination_widget_page_size_50").click()
    
             time.sleep(10)  #Pause for a while to give sometime for the api to send the data 
    
             driver.implicitly_wait(10)  #Conditional wait if the data is existing for 10 sec - if not existing go to exception result
             driver.find_element_by_id("resultTable_tree_table_scroll")
             
             df=pd.read_html(driver.find_element_by_id("resultTable_table").get_attribute('outerHTML'))[0]
             df2=pd.read_html(driver.find_element_by_id("resultTable_tree_table_scroll").get_attribute('outerHTML'))[0]
             df_tbl_name = list(df.columns)
    
             global UPS_18Mod_ITB2_EnergyInfo_DF  #Converted to global because we want to store the data into DataBase Later
             UPS_18Mod_ITB2_EnergyInfo_DF = df2.set_axis(df_tbl_name, axis=1, inplace=False)
             
             del(df,df2,df_tbl_name)
         
             driver.refresh()
         
             Time_AfterTest = datetime.now() #Time log after pressing the submit button
             TimeDiff = abs(Time_AfterTest - Time_BeforeTest) #Difference of before and after the test
             result = "Pass"
             
             print(TestName, TimeDiff, result) 
             
             break
    
        except:
            if retry > 2:
                print("The Test Fails in UPS_18Mod_ITB1_EnergyInfo - Please Re Run the Code", 0/0)
            else:
                print("Number of test done is: ", retry)
                continue  

     
def UPS_001_40Mod_ACEnergy_Dataexport():
    UPS_001_40Mod_AC_Export = UPS_001_40Mod_AC_DF.copy()
    UPS_001_40Mod_AC_Export['Start Time'] = pd.to_datetime(UPS_001_40Mod_AC_Export['Start Time'])
    UPS_001_40Mod_AC_Export = UPS_001_40Mod_AC_Export.sort_values(by='Start Time')
    UPS_001_40Mod_AC_Export['Date'] = UPS_001_40Mod_AC_Export['Start Time'].dt.strftime("%d/%m/%Y")
    UPS_001_40Mod_AC_Export['Time'] = UPS_001_40Mod_AC_Export['Start Time'].dt.strftime('%H:%M')
    
    UPS_001_40Mod_AC_Export = UPS_001_40Mod_AC_Export[['Managed Object',	'Management Domain',	
                                                       'Date',	'Time',	'3QF1 electricity energy(L1)(kWh)',	
                                                       '3QF1 electricity energy(L2)(kWh)',	
                                                       '3QF1 electricity energy(L3)(kWh)',	
                                                       '3QF2 electricity energy(L1)(kWh)',	
                                                       '3QF2 electricity energy(L2)(kWh)',	
                                                       '3QF2 electricity energy(L3)(kWh)',	
                                                       '3QF3 electricity energy(L1)(kWh)',	
                                                       '3QF3 electricity energy(L2)(kWh)',	
                                                       '3QF3 electricity energy(L3)(kWh)',	
                                                       '3QF4 electricity energy(L1)(kWh)',	
                                                       '3QF4 electricity energy(L2)(kWh)',	
                                                       '3QF4 electricity energy(L3)(kWh)',	
                                                       '3QF5 electricity energy(L1)(kWh)',	
                                                       '3QF5 electricity energy(L2)(kWh)',	
                                                       '3QF5 electricity energy(L3)(kWh)']]
    
    # wb = load_workbook(FilePath)
    ws = wb['40Modular_UPS-001_ACEnergy']
    
    
    for row in dataframe_to_rows(UPS_001_40Mod_AC_Export, index=False, header=False):
        ws.append(row)
        
    # wb.save(FilePath)
    # wb.close()

def UPS_001_40Mod_ITB1_EnergyInfo_Dataexport():
    UPS_001_40Mod_ITB1_EnergyInfo_Export = UPS_001_40Mod_ITB1_EnergyInfo_DF.copy()
    UPS_001_40Mod_ITB1_EnergyInfo_Export['Start Time'] = pd.to_datetime(UPS_001_40Mod_ITB1_EnergyInfo_Export['Start Time'])
    UPS_001_40Mod_ITB1_EnergyInfo_Export = UPS_001_40Mod_ITB1_EnergyInfo_Export.sort_values(by='Start Time')
    UPS_001_40Mod_ITB1_EnergyInfo_Export['Date'] = UPS_001_40Mod_ITB1_EnergyInfo_Export['Start Time'].dt.strftime("%d/%m/%Y")
    UPS_001_40Mod_ITB1_EnergyInfo_Export['Time'] = UPS_001_40Mod_ITB1_EnergyInfo_Export['Start Time'].dt.strftime('%H:%M')
    
    UPS_001_40Mod_ITB1_EnergyInfo_Export = UPS_001_40Mod_ITB1_EnergyInfo_Export[['Managed Object',	
                                                                                 'Management Domain',	
                                                                                 'Date',	'Time',	
                                                                                 '1QF1 electricity energy(L1)(kWh)',	
                                                                                 '1QF2 electricity energy(L2)(kWh)',	
                                                                                 '1QF3 electricity energy(L3)(kWh)',	
                                                                                 '1QF4 electricity energy(L1)(kWh)',	
                                                                                 '1QF5 electricity energy(L2)(kWh)',	
                                                                                 '1QF6 electricity energy(L3)(kWh)',	
                                                                                 '1QF7 electricity energy(L1)(kWh)',	
                                                                                 '1QF8 electricity energy(L2)(kWh)',	
                                                                                 '1QF9 electricity energy(L3)(kWh)',	
                                                                                 '1QF10 electricity energy(L1)(kWh)',	
                                                                                 '1QF11 electricity energy(L2)(kWh)',	
                                                                                 '1QF12 electricity energy(L3)(kWh)',	
                                                                                 '1QF13 electricity energy(L1)(kWh)',	
                                                                                 '1QF14 electricity energy(L2)(kWh)',	
                                                                                 '1QF15 electricity energy(L3)(kWh)',	
                                                                                 '1QF16 electricity energy(L1)(kWh)',	
                                                                                 '1QF17 electricity energy(L2)(kWh)',	
                                                                                 '1QF18 electricity energy(L3)(kWh)',	
                                                                                 '1QF19 electricity energy(L1)(kWh)',	
                                                                                 '1QF20 electricity energy(L2)(kWh)',	
                                                                                 '1QF21 electricity energy(L3)(kWh)',	
                                                                                 '1QF22 electricity energy(L1)(kWh)',	
                                                                                 '1QF23 electricity energy(L2)(kWh)',	
                                                                                 '1QF24 electricity energy(L3)(kWh)']]
    
    # wb = load_workbook(FilePath)
    ws = wb['40Modul_UPS-001_ITB1_EnergyInfo']
    
    
    for row in dataframe_to_rows(UPS_001_40Mod_ITB1_EnergyInfo_Export, index=False, header=False):
        ws.append(row)
        
    # wb.save(FilePath)
    # wb.close()

def UPS_001_40Mod_ITB2_EnergyInfo_Dataexport():
    UPS_001_40Mod_ITB2_EnergyInfo_Export = UPS_001_40Mod_ITB2_EnergyInfo_DF.copy()
    UPS_001_40Mod_ITB2_EnergyInfo_Export['Start Time'] = pd.to_datetime(UPS_001_40Mod_ITB2_EnergyInfo_Export['Start Time'])
    UPS_001_40Mod_ITB2_EnergyInfo_Export = UPS_001_40Mod_ITB2_EnergyInfo_Export.sort_values(by='Start Time')
    UPS_001_40Mod_ITB2_EnergyInfo_Export['Date'] = UPS_001_40Mod_ITB2_EnergyInfo_Export['Start Time'].dt.strftime("%d/%m/%Y")
    UPS_001_40Mod_ITB2_EnergyInfo_Export['Time'] = UPS_001_40Mod_ITB2_EnergyInfo_Export['Start Time'].dt.strftime('%H:%M')
    
    UPS_001_40Mod_ITB2_EnergyInfo_Export = UPS_001_40Mod_ITB2_EnergyInfo_Export[['Managed Object',	'Management Domain',	'Date',	'Time',	
                                                                                '2QF1 electricity energy(L1)(kWh)',	
                                                                                '2QF2 electricity energy(L2)(kWh)',	
                                                                                '2QF3 electricity energy(L3)(kWh)',	
                                                                                '2QF4 electricity energy(L1)(kWh)',	
                                                                                '2QF5 electricity energy(L2)(kWh)',	
                                                                                '2QF6 electricity energy(L3)(kWh)',	
                                                                                '2QF7 electricity energy(L1)(kWh)',	
                                                                                '2QF8 electricity energy(L2)(kWh)',	
                                                                                '2QF9 electricity energy(L3)(kWh)',	
                                                                                '2QF10 electricity energy(L1)(kWh)',	
                                                                                '2QF11 electricity energy(L2)(kWh)']]
    # wb = load_workbook(FilePath)
    ws = wb['40Modul_UPS-001_ITB2_EnergyInfo']
    
    
    for row in dataframe_to_rows(UPS_001_40Mod_ITB2_EnergyInfo_Export, index=False, header=False):
        ws.append(row)
        
    # wb.save(FilePath)
    # wb.close()    
    
def UPS_02_40Mod_ITB1_EnergyInfo_Dataexport():
    UPS_02_40Mod_ITB1_EnergyInfo_Export = UPS_02_40Mod_ITB1_EnergyInfo_DF.copy()
    UPS_02_40Mod_ITB1_EnergyInfo_Export['Start Time'] = pd.to_datetime(UPS_02_40Mod_ITB1_EnergyInfo_Export['Start Time'])
    UPS_02_40Mod_ITB1_EnergyInfo_Export = UPS_02_40Mod_ITB1_EnergyInfo_Export.sort_values(by='Start Time')
    UPS_02_40Mod_ITB1_EnergyInfo_Export['Date'] = UPS_02_40Mod_ITB1_EnergyInfo_Export['Start Time'].dt.strftime("%d/%m/%Y")
    UPS_02_40Mod_ITB1_EnergyInfo_Export['Time'] = UPS_02_40Mod_ITB1_EnergyInfo_Export['Start Time'].dt.strftime('%H:%M')
    
    UPS_02_40Mod_ITB1_EnergyInfo_Export = UPS_02_40Mod_ITB1_EnergyInfo_Export[['Managed Object',	'Management Domain',	'Date',	'Time',	
                                                                                '1QF1 electricity energy(L1)(kWh)',	
                                                                                '1QF2 electricity energy(L2)(kWh)',	
                                                                                '1QF3 electricity energy(L3)(kWh)',	
                                                                                '1QF4 electricity energy(L1)(kWh)',	
                                                                                '1QF5 electricity energy(L2)(kWh)',	
                                                                                '1QF6 electricity energy(L3)(kWh)',	
                                                                                '1QF7 electricity energy(L1)(kWh)',	
                                                                                '1QF8 electricity energy(L2)(kWh)',	
                                                                                '1QF9 electricity energy(L3)(kWh)',	
                                                                                '1QF10 electricity energy(L1)(kWh)',	
                                                                                '1QF11 electricity energy(L2)(kWh)',	
                                                                                '1QF12 electricity energy(L3)(kWh)',	
                                                                                '1QF13 electricity energy(L1)(kWh)',	
                                                                                '1QF14 electricity energy(L2)(kWh)',	
                                                                                '1QF15 electricity energy(L3)(kWh)',	
                                                                                '1QF16 electricity energy(L1)(kWh)',	
                                                                                '1QF17 electricity energy(L2)(kWh)',	
                                                                                '1QF18 electricity energy(L3)(kWh)',	
                                                                                '1QF19 electricity energy(L1)(kWh)',	
                                                                                '1QF20 electricity energy(L2)(kWh)',	
                                                                                '1QF21 electricity energy(L3)(kWh)',	
                                                                                '1QF22 electricity energy(L1)(kWh)',	
                                                                                '1QF23 electricity energy(L2)(kWh)',	
                                                                                '1QF24 electricity energy(L3)(kWh)']]
    
    # wb = load_workbook(FilePath)
    ws = wb['40Modul_UPS-02_ITB1_EnergyInfo']
    
    
    for row in dataframe_to_rows(UPS_02_40Mod_ITB1_EnergyInfo_Export, index=False, header=False):
        ws.append(row)
        
    # wb.save(FilePath)
    # wb.close()   

def UPS_02_40Mod_ITB2_EnergyInfo_Dataexport():
    UPS_02_40Mod_ITB2_EnergyInfo_Export = UPS_02_40Mod_ITB2_EnergyInfo_DF.copy()
    UPS_02_40Mod_ITB2_EnergyInfo_Export['Start Time'] = pd.to_datetime(UPS_02_40Mod_ITB2_EnergyInfo_Export['Start Time'])
    UPS_02_40Mod_ITB2_EnergyInfo_Export = UPS_02_40Mod_ITB2_EnergyInfo_Export.sort_values(by='Start Time')
    UPS_02_40Mod_ITB2_EnergyInfo_Export['Date'] = UPS_02_40Mod_ITB2_EnergyInfo_Export['Start Time'].dt.strftime("%d/%m/%Y")
    UPS_02_40Mod_ITB2_EnergyInfo_Export['Time'] = UPS_02_40Mod_ITB2_EnergyInfo_Export['Start Time'].dt.strftime('%H:%M')
    
    UPS_02_40Mod_ITB2_EnergyInfo_Export = UPS_02_40Mod_ITB2_EnergyInfo_Export[[	'Managed Object', 
                                                                            	'Management Domain', 
                                                                            	'Date', 
                                                                            	'Time', 
                                                                            	'2QF1 electricity energy(L1)(kWh)', 
                                                                            	'2QF2 electricity energy(L2)(kWh)', 
                                                                            	'2QF3 electricity energy(L3)(kWh)', 
                                                                            	'2QF4 electricity energy(L1)(kWh)', 
                                                                            	'2QF5 electricity energy(L2)(kWh)', 
                                                                            	'2QF6 electricity energy(L3)(kWh)', 
                                                                            	'2QF7 electricity energy(L1)(kWh)', 
                                                                            	'2QF8 electricity energy(L2)(kWh)', 
                                                                            	'2QF9 electricity energy(L3)(kWh)', 
                                                                            	'2QF10 electricity energy(L1)(kWh)', 
                                                                            	'2QF11 electricity energy(L2)(kWh)']]
    
    # wb = load_workbook(FilePath)
    ws = wb['40Modul_UPS-02_ITB2_EnergyInfo']
    
    
    for row in dataframe_to_rows(UPS_02_40Mod_ITB2_EnergyInfo_Export, index=False, header=False):
        ws.append(row)
        
    # wb.save(FilePath)
    # wb.close()       

def UPS_18Mod_ACEnergy_Dataexport():
    UPS_18Mod_ACEnergy_Export = UPS_18Mod_ACEnergy_DF.copy()
    UPS_18Mod_ACEnergy_Export['Start Time'] = pd.to_datetime(UPS_18Mod_ACEnergy_Export['Start Time'])
    UPS_18Mod_ACEnergy_Export = UPS_18Mod_ACEnergy_Export.sort_values(by='Start Time')
    UPS_18Mod_ACEnergy_Export['Date'] = UPS_18Mod_ACEnergy_Export['Start Time'].dt.strftime("%d/%m/%Y")
    UPS_18Mod_ACEnergy_Export['Time'] = UPS_18Mod_ACEnergy_Export['Start Time'].dt.strftime('%H:%M')
    
    UPS_18Mod_ACEnergy_Export = UPS_18Mod_ACEnergy_Export[['Managed Object', 
                                                        	'Management Domain', 
                                                        	'Date', 
                                                        	'Time',
                                                        	'3QF1 electricity energy(L1)(kWh)', 
                                                        	'3QF1 electricity energy(L2)(kWh)', 
                                                        	'3QF1 electricity energy(L3)(kWh)', 
                                                        	'3QF2 electricity energy(L1)(kWh)', 
                                                        	'3QF2 electricity energy(L2)(kWh)', 
                                                        	'3QF2 electricity energy(L3)(kWh)']]
    
    # wb = load_workbook(FilePath)
    ws = wb['18Modular_UPS_ACEnergy']
    
    for row in dataframe_to_rows(UPS_18Mod_ACEnergy_Export, index=False, header=False):
        ws.append(row)
        
    # wb.save(FilePath)
    # wb.close()       

def UPS_18Mod_ITB1_EnergyInfo_Dataexport():
    UPS_18Mod_ITB1_EnergyInfo_Export = UPS_18Mod_ITB1_EnergyInfo_DF.copy()
    UPS_18Mod_ITB1_EnergyInfo_Export['Start Time'] = pd.to_datetime(UPS_18Mod_ITB1_EnergyInfo_Export['Start Time'])
    UPS_18Mod_ITB1_EnergyInfo_Export = UPS_18Mod_ITB1_EnergyInfo_Export.sort_values(by='Start Time')
    UPS_18Mod_ITB1_EnergyInfo_Export['Date'] = UPS_18Mod_ITB1_EnergyInfo_Export['Start Time'].dt.strftime("%d/%m/%Y")
    UPS_18Mod_ITB1_EnergyInfo_Export['Time'] = UPS_18Mod_ITB1_EnergyInfo_Export['Start Time'].dt.strftime('%H:%M')
    
    UPS_18Mod_ITB1_EnergyInfo_Export = UPS_18Mod_ITB1_EnergyInfo_Export[['Managed Object',	'Management Domain',	'Date',	'Time',	
                                                                    	'1QF1 electricity energy(L1)(kWh)', 
                                                                    	'1QF2 electricity energy(L2)(kWh)', 
                                                                    	'1QF3 electricity energy(L3)(kWh)', 
                                                                    	'1QF4 electricity energy(L1)(kWh)', 
                                                                    	'1QF5 electricity energy(L2)(kWh)', 
                                                                    	'1QF6 electricity energy(L3)(kWh)', 
                                                                    	'1QF7 electricity energy(L1)(kWh)', 
                                                                    	'1QF8 electricity energy(L2)(kWh)', 
                                                                    	'1QF9 electricity energy(L3)(kWh)', 
                                                                    	'1QF10 electricity energy(L1)(kWh)', 
                                                                    	'1QF11 electricity energy(L2)(kWh)', 
                                                                    	'1QF12 electricity energy(L3)(kWh)', 
                                                                    	'1QF13 electricity energy(L1)(kWh)', 
                                                                    	'1QF14 electricity energy(L2)(kWh)', 
                                                                    	'1QF15 electricity energy(L3)(kWh)', 
                                                                    	'1QF16 electricity energy(L1)(kWh)', 
                                                                    	'1QF17 electricity energy(L2)(kWh)', 
                                                                    	'1QF18 electricity energy(L3)(kWh)', 
                                                                    	'1QF19 electricity energy(L1)(kWh)', 
                                                                    	'1QF20 electricity energy(L2)(kWh)', 
                                                                    	'1QF21 electricity energy(L3)(kWh)']]
    
    # wb = load_workbook(FilePath)
    ws = wb['18Modul_UPS_ITB1_EnergyInfo']
    
    
    for row in dataframe_to_rows(UPS_18Mod_ITB1_EnergyInfo_Export, index=False, header=False):
        ws.append(row)
        
    # wb.save(FilePath)
    # wb.close()       

def UPS_18Mod_ITB2_EnergyInfo_Dataexport():
   UPS_18Mod_ITB2_EnergyInfo_Export = UPS_18Mod_ITB2_EnergyInfo_DF.copy()
   UPS_18Mod_ITB2_EnergyInfo_Export['Start Time'] = pd.to_datetime(UPS_18Mod_ITB2_EnergyInfo_Export['Start Time'])
   UPS_18Mod_ITB2_EnergyInfo_Export = UPS_18Mod_ITB2_EnergyInfo_Export.sort_values(by='Start Time')
   UPS_18Mod_ITB2_EnergyInfo_Export['Date'] = UPS_18Mod_ITB2_EnergyInfo_Export['Start Time'].dt.strftime("%d/%m/%Y")
   UPS_18Mod_ITB2_EnergyInfo_Export['Time'] = UPS_18Mod_ITB2_EnergyInfo_Export['Start Time'].dt.strftime('%H:%M')
   
   UPS_18Mod_ITB2_EnergyInfo_Export = UPS_18Mod_ITB2_EnergyInfo_Export[['Managed Object',	'Management Domain',	'Date',	'Time',	
                                                                    	'2QF1 electricity energy(L1)(kWh)', 
                                                                    	'2QF2 electricity energy(L2)(kWh)', 
                                                                    	'2QF3 electricity energy(L3)(kWh)', 
                                                                    	'2QF4 electricity energy(L1)(kWh)', 
                                                                    	'2QF5 electricity energy(L2)(kWh)', 
                                                                    	'2QF6 electricity energy(L3)(kWh)', 
                                                                    	'2QF7 electricity energy(L1)(kWh)', 
                                                                    	'2QF8 electricity energy(L2)(kWh)', 
                                                                    	'2QF9 electricity energy(L3)(kWh)', 
                                                                    	'2QF10 electricity energy(L1)(kWh)', 
                                                                    	'2QF11 electricity energy(L2)(kWh)', 
                                                                    	'2QF12 electricity energy(L3)(kWh)', 
                                                                    	'2QF13 electricity energy(L1)(kWh)', 
                                                                    	'2QF14 electricity energy(L2)(kWh)', 
                                                                    	'2QF15 electricity energy(L3)(kWh)', 
                                                                    	'2QF16 electricity energy(L1)(kWh)', 
                                                                    	'2QF17 electricity energy(L2)(kWh)', 
                                                                    	'2QF18 electricity energy(L3)(kWh)', 
                                                                    	'2QF19 electricity energy(L1)(kWh)', 
                                                                    	'2QF20 electricity energy(L2)(kWh)', 
                                                                    	'2QF21 electricity energy(L3)(kWh)']]
   
   # wb = load_workbook(FilePath)
   ws = wb['18Modul_UPS_ITB2_EnergyInfo']
   
   
   for row in dataframe_to_rows(UPS_18Mod_ITB2_EnergyInfo_Export, index=False, header=False):
       ws.append(row)
       
   # wb.save(FilePath)
   # wb.close()        

def Powerquery_Update():
    xl = win32com.client.DispatchEx("Excel.Application")
    wb = xl.workbooks.open(r'F:\Shared\Ashner\DataCenter_PowerMonitoring\Datacenter_Power_Test\MEOL_DataCenter_PowerMonitoring_2022.xlsx')
    xl.Visible = True
    wb.RefreshAll()
    time.sleep(120)
    wb.Close(True)
    xl.Quit()

if __name__ == '__main__':
    
    time_start = datetime.now()
    
    #1. Need to load the web page of Net Eco
    NetEco_Webloading()
    
    #2. Login to the NetEco
    NetEco_Login()
    
    #2.1 If License Near Expiration Found it will run this code and click, if not found it will break
    License_Near_Expiry()
    
    #3. Historical Report Page
    Historical_Report()
     
    'In this test Datalogging the 40 Modular UPS_001 AC Energy'
    #4. Set the time scope of data to be captured
    Set_Duration_Data()
    #5. Expand till DataCenter01 (It means 40 Modular and 18 Modular will apper)
    Expand_Openlab_DC_Monitoring()
    #6. Get the Data for 40Modular AC Energy (Refresh will also happen on this function)
    UPS_001_40Mod_ACEnergy()
    
    'Pause for a while transition to next Data Web Scrapping'
    time.sleep(3)
    
    'In this test Datalogging the 40 Modular UPS_001 ITB1 Energy Measurement'
    #7. Set the time scope of data to be captured
    Set_Duration_Data()
    #8. Expand till DataCenter01 (It means 40 Modular and 18 Modular will apper)
    Expand_Openlab_DC_Monitoring()
    #9 Get the Data for 40Modular ITB1 Energy Measure (Refresh will also happen on this function)
    UPS_001_40Mod_ITB1_EnergyInfo()
    
    'Pause for a while transition to next Data Web Scrapping'
    time.sleep(3)
    
    'In this test Datalogging the 40 Modular UPS_001 ITB2 Energy Measurement'
    #10. Set the time scope of data to be captured
    Set_Duration_Data()
    #11. Expand till DataCenter01 (It means 40 Modular and 18 Modular will apper)
    Expand_Openlab_DC_Monitoring()
    #12. Get the Data for 40Modular ITB2 Energy Measure (Refresh will also happen on this function)
    UPS_001_40Mod_ITB2_EnergyInfo()  
    
    'Pause for a while transition to next Data Web Scrapping'
    time.sleep(3)
    
    'In this test Datalogging the 40 Modular UPS_02 ITB1 Energy Measurement'
    #13. Set the time scope of data to be captured
    Set_Duration_Data()
    #14. Expand till DataCenter01 (It means 40 Modular and 18 Modular will apper)
    Expand_Openlab_DC_Monitoring()
    #15. Get the Data for 40Modular ITB2 Energy Measure (Refresh will also happen on this function)
    UPS_02_40Mod_ITB1_EnergyInfo()  
    
    'Pause for a while transition to next Data Web Scrapping'
    time.sleep(3)
    
    'In this test Datalogging the 40 Modular UPS_02 ITB2 Energy Measurement'
    #16. Set the time scope of data to be captured
    Set_Duration_Data()
    #17. Expand till DataCenter01 (It means 40 Modular and 18 Modular will apper)
    Expand_Openlab_DC_Monitoring()
    #18. Get the Data for 40Modular ITB2 Energy Measure (Refresh will also happen on this function)
    UPS_02_40Mod_ITB2_EnergyInfo()  
    
    'Pause for a while transition to next Data Web Scrapping'
    time.sleep(3)
    
    'In this test Datalogging the 18 Modular UPS AC Energy'
    #19. Set the time scope of data to be captured
    Set_Duration_Data()
    #20. Expand till DataCenter01 (It means 40 Modular and 18 Modular will apper)
    Expand_Openlab_DC_Monitoring()
    #21. Get the Data for 40Modular ITB2 Energy Measure (Refresh will also happen on this function)
    UPS_18Mod_ACEnergy()      
    
    'Pause for a while transition to next Data Web Scrapping'
    time.sleep(3)
    
    'In this test Datalogging the 18 Modular UPS ITB1 Energy Measurement'
    #22. Set the time scope of data to be captured
    Set_Duration_Data()
    #23. Expand till DataCenter01 (It means 40 Modular and 18 Modular will apper)
    Expand_Openlab_DC_Monitoring()
    #24. Get the Data for 40Modular ITB2 Energy Measure (Refresh will also happen on this function)
    UPS_18Mod_ITB1_EnergyInfo() 
    
    'Pause for a while transition to next Data Web Scrapping'
    time.sleep(3)
    
    'In this test Datalogging the 18 Modular UPS ITB1 Energy Measurement'
    #25. Set the time scope of data to be captured
    Set_Duration_Data()
    #26. Expand till DataCenter01 (It means 40 Modular and 18 Modular will apper)
    Expand_Openlab_DC_Monitoring()
    #27. Get the Data for 40Modular ITB2 Energy Measure (Refresh will also happen on this function)
    UPS_18Mod_ITB2_EnergyInfo() 
    
    time.sleep(3)

    driver.quit()
    
    time_end = datetime.now()
    
    time_difference = abs(time_end-time_start)
    
    print(time_difference)
    
    
    #This part is datalogging of Data Center Energy Information
    FilePath = r'F:\Shared\Ashner\DataCenter_PowerMonitoring\Datacenter_Power_Test\MEOL_DataCenter_PowerMonitoring_Reference.xlsx'

    wb = load_workbook(FilePath)
    
    UPS_001_40Mod_ACEnergy_Dataexport()
    time.sleep(2)
    UPS_001_40Mod_ITB1_EnergyInfo_Dataexport()
    time.sleep(2)
    UPS_001_40Mod_ITB2_EnergyInfo_Dataexport()
    time.sleep(2)
    UPS_02_40Mod_ITB1_EnergyInfo_Dataexport()
    time.sleep(2)
    UPS_02_40Mod_ITB2_EnergyInfo_Dataexport()
    time.sleep(2)
    UPS_18Mod_ACEnergy_Dataexport()
    time.sleep(2)
    UPS_18Mod_ITB1_EnergyInfo_Dataexport()
    time.sleep(2)
    UPS_18Mod_ITB2_EnergyInfo_Dataexport()
    
    wb.save(FilePath)
    wb.close() 
    
    #Refresh the PowerQuery. This connects to the formula side of the data
    Powerquery_Update()
    